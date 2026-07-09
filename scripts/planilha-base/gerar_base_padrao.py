#!/usr/bin/env python3
"""Gerador da BASE-PADRAO-v1.xlsx — planilha-base de entrada de dados da obra.

Espelha docs/rma/TEMPLATE-PLANILHA-BASE.md (Fase 2 · layout final, decisões 03/jul/2026):
3 abas meta (cinza) + 18 imutáveis I-nn (azul) + 13 mutáveis M-nn (verde).
Regras do Contrato do Template (FLUXO-UPLOAD-E-TEMPLATE.md §5): 1 tabela lógica por aba,
sem linha-subtítulo (agrupar = coluna Grupo), cabeçalho estável, TOTAL de controle,
código de seção no título.

Uso:  python3 scripts/planilha-base/gerar_base_padrao.py
Saída: docs/rma/templates/BASE-PADRAO-v1.xlsx
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = Path(__file__).resolve().parents[2]
SAIDA = RAIZ / "docs" / "rma" / "templates" / "BASE-PADRAO-v1.xlsx"

# ── paleta / estilos (espelha tokens do DS: ink · brand · semânticos) ─────────
COR_META = "6B7280"
COR_IMUT = "1F4E79"  # ink/navy — baseline contratual
COR_MUT = "2E7D32"  # verde — ciclo mensal
F_TITULO = Font(name="Calibri", size=14, bold=True, color="1F4E79")
F_INSTR = Font(name="Calibri", size=9, italic=True, color="6B7280")
F_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
F_TOTAL = Font(name="Calibri", size=10, bold=True)
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_HEADER_MUT = PatternFill("solid", fgColor="2E7D32")
FILL_TOTAL = PatternFill("solid", fgColor="E8EDF3")
FILL_CAMPO = PatternFill("solid", fgColor="F3F4F6")
BORDA = Border(*[Side(style="thin", color="C9CFD8")] * 4)
ALINHA_H = Alignment(horizontal="center", vertical="center", wrap_text=True)

FILL_ZEBRA = PatternFill("solid", fgColor="F4F7FB")
FILL_INPUT = PatternFill("solid", fgColor="FFFDF2")  # célula de entrada (chave-valor)
FILL_DECL = PatternFill("solid", fgColor="FFE9B8")  # TOTAL DECLARADO — "digite aqui"
F_TITULO_M = Font(name="Calibri", size=14, bold=True, color="2E7D32")
F_CHIP_I = Font(name="Calibri", size=9, bold=True, color="1F4E79")
F_CHIP_M = Font(name="Calibri", size=9, bold=True, color="2E7D32")

FMT_RS = '#,##0.00'
FMT_PCT = "0.00%"
FMT_DATA = "DD/MM/YYYY"
FMT_INT = "#,##0"

N_MESES = 60  # colunas das matrizes wide (obra de até 5 anos; apagar sobras)


def aba(wb, nome, cor):
    ws = wb.create_sheet(nome)
    ws.sheet_properties.tabColor = cor
    return ws


def titulo(ws, texto, instrucao, mut=False):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    ws.merge_cells("A3:H3")
    ws["A1"] = texto
    ws["A1"].font = F_TITULO_M if mut else F_TITULO
    ws["A2"] = instrucao
    ws["A2"].font = F_INSTR
    ws["A2"].alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 30
    if ws.title.startswith("I-"):
        ws["A3"] = "● PARTE 1 · IMUTÁVEL — preencher 1× no onboarding (muda só por aditivo/TAC)"
        ws["A3"].font = F_CHIP_I
    elif ws.title.startswith("M-"):
        ws["A3"] = "● PARTE 2 · MENSAL — preencher a cada BM (ciclos ACUMULAM · nunca sobrescrever)"
        ws["A3"].font = F_CHIP_M
    ws.freeze_panes = "A5"


def tabela(ws, headers, widths, n_rows, fmts=None, total_sum_cols=(), mut=False, header_row=4, filtro=True):
    """Tabela única da aba: header estilizado + n_rows vazias + TOTAL opcional.

    fmts: dict col_idx(1-based)->number_format aplicado às linhas de dado.
    total_sum_cols: colunas (1-based) que ganham =SUM() na linha TOTAL.
    Retorna (primeira_linha_dado, ultima_linha_dado).
    """
    fill = FILL_HEADER_MUT if mut else FILL_HEADER
    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=header_row, column=j, value=h)
        c.font, c.fill, c.alignment, c.border = F_HEADER, fill, ALINHA_H, BORDA
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[header_row].height = 30
    r0, r1 = header_row + 1, header_row + n_rows
    for r in range(r0, r1 + 1):
        zebra = (r - r0) % 2 == 1
        for j in range(1, len(headers) + 1):
            c = ws.cell(row=r, column=j)
            c.border = BORDA
            if zebra:
                c.fill = FILL_ZEBRA
            if fmts and j in fmts:
                c.number_format = fmts[j]
    if filtro and n_rows > 0:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{r1}"
    if total_sum_cols:
        rt = r1 + 1
        c = ws.cell(row=rt, column=1, value="TOTAL (controle · não apagar)")
        c.font, c.fill = F_TOTAL, FILL_TOTAL
        for j in range(1, len(headers) + 1):
            ws.cell(row=rt, column=j).fill = FILL_TOTAL
            ws.cell(row=rt, column=j).border = BORDA
        for j in total_sum_cols:
            L = get_column_letter(j)
            c = ws.cell(row=rt, column=j, value=f"=SUM({L}{r0}:{L}{r1})")
            c.font, c.number_format = F_TOTAL, (fmts or {}).get(j, FMT_RS)
    return r0, r1


def chave_valor(ws, campos, header_row=4):
    """Tabela Campo | Valor | Observação. campos = [(campo, fmt|None, obs)]."""
    for j, (h, w) in enumerate(zip(["Campo", "Valor", "Observação"], [42, 30, 52]), start=1):
        c = ws.cell(row=header_row, column=j, value=h)
        c.font, c.fill, c.alignment, c.border = F_HEADER, FILL_HEADER, ALINHA_H, BORDA
        ws.column_dimensions[get_column_letter(j)].width = w
    for i, (campo, fmt, obs) in enumerate(campos):
        r = header_row + 1 + i
        a = ws.cell(row=r, column=1, value=campo)
        a.fill, a.border = FILL_CAMPO, BORDA
        v = ws.cell(row=r, column=2)
        v.border = BORDA
        v.fill = FILL_INPUT
        if fmt:
            v.number_format = fmt
        o = ws.cell(row=r, column=3, value=obs)
        o.font, o.border = F_INSTR, BORDA




def total_declarado(ws, row, value_col, fmt, nota):
    """Linha de total DECLARADO (manual): o motor confere Σ das linhas não-subtotal contra ele."""
    c = ws.cell(row=row, column=1, value="TOTAL DECLARADO (digite o total oficial — o motor confere)")
    c.font, c.fill = F_TOTAL, FILL_TOTAL
    for j in range(1, value_col + 1):
        ws.cell(row=row, column=j).fill = FILL_TOTAL
        ws.cell(row=row, column=j).border = BORDA
    v = ws.cell(row=row, column=value_col)
    v.number_format, v.font, v.fill = fmt, F_TOTAL, FILL_DECL
    o = ws.cell(row=row, column=value_col + 1, value=nota)
    o.font = F_INSTR


def dv_lista(ws, ref_listas, celulas):
    dv = DataValidation(type="list", formula1=ref_listas, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(celulas)


def matriz_meses(ws, primeira_col_fixa, fixas, widths_fixas, n_rows, mut=False, fmt=FMT_RS):
    """Matriz wide: colunas fixas + Mês 1..N_MESES + TOTAL por linha."""
    headers = list(fixas) + [str(m) for m in range(1, N_MESES + 1)] + ["TOTAL linha"]
    widths = list(widths_fixas) + [11] * N_MESES + [14]
    fmts = {j: fmt for j in range(len(fixas) + 1, len(headers) + 1)}
    r0, r1 = tabela(ws, headers, widths, n_rows, fmts=fmts, mut=mut, filtro=False)
    ws.freeze_panes = f"{get_column_letter(len(fixas) + 1)}5"  # fixas visíveis no scroll horizontal
    c0 = get_column_letter(len(fixas) + 1)
    c1 = get_column_letter(len(fixas) + N_MESES)
    for r in range(r0, r1 + 1):
        cell = ws.cell(row=r, column=len(headers), value=f"=SUM({c0}{r}:{c1}{r})")
        cell.number_format = fmt
        cell.font = F_TOTAL
    _ = primeira_col_fixa
    return r0, r1


# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)

# ── LEIA-ME ───────────────────────────────────────────────────────────────────
ws = aba(wb, "LEIA-ME", COR_META)
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 118
ws.row_dimensions[1].height = 34
linhas = [
    ("BASE-PADRAO v1 · Planilha-Base da Obra — Plataforma de Administração Contratual IA", F_TITULO),
    ("", None),
    ("COMO USAR", F_TOTAL),
    ("· PARTE 1 — abas AZUIS (I-01..I-18): dados IMUTÁVEIS do contrato. Preencha 1× no onboarding.", None),
    ("  Só mudam por aditivo/TAC — e aí é NOVA VERSÃO do arquivo (nunca edição silenciosa).", None),
    ("· PARTE 2 — abas VERDES (M-01..M-13): ciclo MENSAL. Preencha a cada Boletim de Medição (BM).", None),
    ("· Abas CINZAS: meta (este guia, Guia da IA, LISTAS dos dropdowns). Não preencher dado de obra.", None),
    ("", None),
    ("REGRAS DE OURO (o motor confere tudo — planilha auto-verificável)", F_TOTAL),
    ("1. NUNCA digite dado derivado: acumulados, %, desvios, faróis, desequilíbrios. O motor recompute TUDO.", None),
    ("2. PENDENTE ≠ ZERO: o que ainda não foi medido fica EM BRANCO (nunca 0). Branco = honesto.", None),
    ("3. NÃO mexa nas linhas TOTAL (controle de conservação) nem nos cabeçalhos das colunas.", None),
    ("4. Não insira linhas-subtítulo no meio das tabelas — para agrupar, use a coluna Grupo.", None),
    ("5. Salve o arquivo COM OS VALORES calculados (Excel normal salva; evite exportadores exóticos).", None),
    ("6. O faturamento REAL não se digita: ele DERIVA da Medição do BM (M-02) × PQ (I-02).", None),
    ("7. Insumos de Faturamento Direto: marque na PQ (coluna 'Insumo FD?') — não há aba separada.", None),
    ("8. Matrizes mensais têm 60 colunas de mês: use as que a obra precisar e deixe o resto em branco.", None),
    ("9. CICLOS ACUMULAM: nas abas mensais, cada BM ganha linhas NOVAS (coluna BM nº / Competência).", None),
    ("   NUNCA apague nem sobrescreva um ciclo anterior — a planilha é o histórico da obra inteira.", None),
    ("10. Percentuais no formato % do Excel (digite 5% ou 0,05 — nunca 5 querendo dizer 5%).", None),
    ("11. Precisa de mais linhas? Insira DENTRO do intervalo (acima da linha TOTAL) — as fórmulas expandem.", None),
    ("    Obra com mais de 60 meses: insira colunas ANTES da última coluna de mês (o TOTAL linha expande).", None),
    ("12. A aba LISTAS é DA OBRA: edite disciplinas/tipos para o seu tipo de obra (rodovia é só exemplo).", None),
    ("13. Aba que não se aplica ao tipo da obra? Marque Aplicável = Não no Guia da IA (não deixe pela metade).", None),
    ("", None),
    ("DÚVIDAS DE PREENCHIMENTO → docs/rma/TEMPLATE-PLANILHA-BASE.md (mapa completo campo a campo).", F_INSTR),
]
for i, (txt, f) in enumerate(linhas, start=1):
    c = ws.cell(row=i, column=1, value=txt)
    c.font = f or Font(name="Calibri", size=10)
    if i == 1:  # banner
        c.font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(vertical="center", indent=1)
    elif txt.startswith("· PARTE 1"):
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
    elif txt.startswith("· PARTE 2"):
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E7D32")
    elif txt in ("COMO USAR", "REGRAS DE OURO (o motor confere tudo — planilha auto-verificável)"):
        c.fill = PatternFill("solid", fgColor="E8EDF3")

# ── LISTAS (fonte dos dropdowns) ──────────────────────────────────────────────
ws = aba(wb, "LISTAS", COR_META)
ws["A25"] = "⚠ Estas listas são DA OBRA: edite/estenda para o seu tipo de obra (rodovia é só o exemplo pré-carregado)."
ws["A25"].font = F_INSTR
LISTAS = {
    "A": ("Disciplina", ["Terraplenagem", "Pavimentação", "Drenagem", "OAE / Pontes / Dispositivos",
                          "Obras Complementares", "Sinalização", "Serviços Preliminares", "Meio Ambiente",
                          "Contenções", "Administração Local", "Mobiliz./Desmobiliz.", "Geodrenos",
                          "Recuperação de Sinistros", "Insumos (Fat. Direto)", "Outros"]),
    "B": ("Categoria", ["MOD", "MOI", "EQP"]),
    "C": ("Fonte de índice", ["SINAPI", "DNIT", "ANP", "SBC", "EMOP", "SCO", "OUTRA"]),
    "D": ("Tipo de fonte", ["índice", "preço"]),
    "E": ("Sim/Não", ["Sim", "Não"]),
    "F": ("Tipo de segmento", ["Duplicação", "Restauração", "Implantação", "Sinistro", "Outro"]),
    "G": ("Categoria de evento", ["Chuva", "Impedimento", "Paralisação", "Interferência", "Pleito", "Outro"]),
    "H": ("Tipo de elemento", ["OAE", "Dispositivo", "Talude", "Outro"]),
    "I": ("Cenário M1", ["cpus", "proposta", "assinatura", ""]),
    "J": ("UF", ["AC", "AL", "AM", "AP", "BA", "CE", "DF", "ES", "GO", "MA", "MG", "MS", "MT", "PA",
                  "PB", "PE", "PI", "PR", "RJ", "RN", "RO", "RR", "RS", "SC", "SE", "SP", "TO"]),
}
REF = {}
ws.sheet_view.showGridLines = False
for col, (nome, vals) in LISTAS.items():
    ws[f"{col}1"] = nome
    ws[f"{col}1"].font = F_HEADER
    ws[f"{col}1"].fill = FILL_HEADER
    ws[f"{col}1"].alignment = ALINHA_H
    for i, v in enumerate(vals, start=2):
        ws[f"{col}{i}"] = v
    ws.column_dimensions[col].width = 24
    # +24 linhas de FOLGA: a obra pode ESTENDER a lista (disciplinas/tipos do seu tipo de obra)
    # e os dropdowns enxergam — blanks são ignorados pelo Excel.
    REF[nome] = f"=LISTAS!${col}$2:${col}${len(vals) + 25}"

# ── Guia da IA (lista autoritativa p/ gate de cobertura) ─────────────────────
ws = aba(wb, "Guia da IA", COR_META)
titulo(ws, "Guia da IA — mapa autoritativo das abas (gate de cobertura)",
       "Toda aba ATÔMICA marcada Aplicável=Sim e não roteada vira pendência explícita. Marque Aplicável=Não nas abas que não valem para o tipo da obra (ex.: mapa por km em obra não-linear; chuva diária sem RDO) — o gate ignora. Não renomear abas sem atualizar este guia.")
GUIA = [
    # (aba, código, classificação, preenchimento, lê de, alimenta)
    ("I-01 Identificação", "C.1", "atômica", "1× (onboarding)", "contrato assinado", "cadastro · C.1 · header de tudo"),
    ("I-02 PQ", "C.0", "atômica", "1×", "Anexo de quantidades", "C.0 · Curva ABC · D.4 VA · rollups disciplina/frente · insumos FD (col Insumo FD?)"),
    ("I-03 BDI", "C.1", "atômica", "1×", "composição do BDI da proposta", "C.1 · D.2"),
    ("I-04 Encargos", "D.3", "atômica", "1×", "composição de encargos da proposta", "D.3"),
    ("I-05 CPUs", "CPU", "atômica", "1×", "composições de custo unitário", "C.7 física · D.4 VA"),
    ("I-06 Curva Prevista", "C.3/C.5", "atômica", "1×", "cronograma fís-fin contratado", "C.3 previsto · C.5 curva física"),
    ("I-07 Previsto R$ Disc x Mês", "C.3", "atômica", "1×", "cronograma fís-fin", "C.3 matriz · C.14 avanço"),
    ("I-08 Previsto Fís Disc x Mês", "C.5", "atômica", "1×", "cronograma físico", "C.5 matriz física"),
    ("I-19 Previsto R$ Frente x Mês", "C.3", "atômica", "1×", "cronograma fís-fin", "C.3 por frente · C.8 curvas por frente"),
    ("I-09 Gantt Tarefas", "C.13", "atômica", "1×", "MS Project / cronograma", "C.13 Timeline"),
    ("I-10 Marcos", "C.5", "atômica", "1×", "contrato (marcos)", "C.5 marcos"),
    ("I-11 Recursos Contratados", "C.4", "atômica", "1×", "histograma da proposta", "C.4 · D.1 (col Grupo Adm Local) · D.3"),
    ("I-12 Histograma Contratado", "C.4", "atômica", "1×", "histograma da proposta", "C.4 mobilização prevista"),
    ("I-13 Fontes de Índice", "C.6", "atômica", "1×", "índices na data da OS", "C.6/D.5 seletor multifonte"),
    ("I-14 Reequilíbrio Params", "D.5", "atômica", "1×", "contrato (cláusulas reajuste)", "D.5 M1/M2"),
    ("I-15 Mapa Segmentos", "C.14", "atômica", "1×", "projeto (km/frentes)", "C.14 · C.8"),
    ("I-16 Mapa Elementos", "C.14", "atômica", "1×", "projeto (OAE/dispositivos)", "C.14 Bloco 5"),
    ("I-17 Chuvas Baseline", "C.9", "atômica", "1×", "série histórica (INMET etc.)", "C.9 · D.6"),
    ("I-18 Parâmetros da Obra", "config", "atômica", "1×", "metas internas", "C.7 metas · D.6 params · faróis"),
    ("M-01 Capa do Ciclo", "BM", "atômica", "mensal", "BM corrente", "corte de TODOS os domínios"),
    ("M-02 Medição BM", "C.3", "atômica", "mensal", "Boletim de Medição", "faturamento real · físico · D.4 qtd medida · C.6 qtd medida"),
    ("M-03 Recursos Reais", "C.4", "atômica", "mensal", "apontamento/folha", "C.4 real · C.7 HH · D.1 real · D.3"),
    ("M-04 Avanço Físico Real", "C.5", "atômica", "mensal", "medição física", "C.5 real · C.14"),
    ("M-05 Gantt Real", "C.13", "atômica", "mensal", "cronograma atualizado", "C.13 barras reais"),
    ("M-06 Marcos Real", "C.5", "atômica", "mensal", "medição", "C.5 marcos %"),
    ("M-07 Chuvas do Mês", "C.9", "atômica", "mensal", "pluviômetro/RDO", "C.9 real · D.6"),
    ("M-08 Chuva Diária RDO", "D.6", "atômica (opcional)", "mensal", "RDO diário", "D.6 dias a cobrar · C.9 headline"),
    ("M-09 Eventos e Impedimentos", "C.13/D.6", "atômica", "mensal (incremental)", "RDO/atas/CEs/ofícios", "C.13 eventos · D.6 · Windows Analysis"),
    ("M-10 Liberação de Frentes", "C.14", "atômica", "mensal", "ordens de liberação", "C.14 real · C.8 liberado"),
    ("M-11 Série Índice Contratual", "D.5", "atômica", "mensal (append)", "IBGE/FGV publicação", "D.5 M1 · reajuste"),
    ("M-12 Índices dos Insumos", "C.6", "atômica", "mensal", "DNIT/SINAPI/ANP/SBC/EMOP/SCO", "C.6/D.5 valor atual das fontes"),
    ("M-13 Custos Pagos", "C.6", "atômica (opcional)", "mensal", "notas fiscais", "C.6 preço real · C.7 custo real"),
    ("— (sem aba)", "D.0–D.11", "derivada", "—", "todas acima", "Painel de Desequilíbrio — o motor RECOMPUTA; nunca digitar"),
]
tabela(ws, ["Aba", "Código de seção", "Classificação", "Aplicável?", "Preenchimento", "Lê de", "Alimenta"],
       [30, 15, 18, 12, 20, 34, 52], 0, header_row=4)
for i, row in enumerate(GUIA):
    r = 5 + i
    valores = list(row[:3]) + ["—" if row[2] == "derivada" else "Sim"] + list(row[3:])
    for j, v in enumerate(valores, start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.border = BORDA
        c.font = Font(name="Calibri", size=9)
        if i % 2 == 1:
            c.fill = FILL_ZEBRA
dv_lista(ws, REF["Sim/Não"], f"D5:D{4 + len(GUIA)}")

# ══════════════ PARTE 1 · IMUTÁVEIS ══════════════════════════════════════════

ws = aba(wb, "I-01 Identificação", COR_IMUT)
titulo(ws, "I-01 · C.1 — Identificação do Contrato",
       "Preencher 1× no onboarding. OS real pode ficar em branco até ser emitida (preencher 1× quando sair).")
chave_valor(ws, [
    ("Nome interno da obra", None, "ex.: BR-101 — Lote 2 Macaé"),
    ("Objeto contratado", None, "descrição do objeto conforme contrato"),
    ("Contratante", None, ""),
    ("Contratada", None, ""),
    ("Modalidade", None, "ex.: empreitada por preço unitário"),
    ("Cidade", None, ""),
    ("UF", None, "sigla — ver aba LISTAS"),
    ("Valor contratual — PV com BDI (R$)", FMT_RS, "preço de venda total do contrato"),
    ("Custo direto (R$)", FMT_RS, "da composição do BDI"),
    ("Custo indireto (R$)", FMT_RS, ""),
    ("Receita / markup (R$)", FMT_RS, "PV − custo total"),
    ("Data de assinatura", FMT_DATA, ""),
    ("OS original (prevista)", FMT_DATA, ""),
    ("OS real (emitida)", FMT_DATA, "deixar EM BRANCO se ainda não emitida"),
    ("Data de início", FMT_DATA, ""),
    ("Término contratual", FMT_DATA, ""),
    ("Prazo (meses)", FMT_INT, ""),
    ("Índice de reajuste", None, "ex.: INCC-DI, IPCA — conforme cláusula"),
    ("Periodicidade do reajuste", None, "ex.: anual (aniversário)"),
    ("Data-base da proposta", FMT_DATA, ""),
    ("Gestor da obra", None, "cargo/função (não usar nome de pessoa física)"),
    ("Adm contratual", None, "cargo/função"),
])
dv_lista(ws, REF["UF"], "B11")

ws = aba(wb, "I-02 PQ", COR_IMUT)
titulo(ws, "I-02 · C.0 — Planilha de Quantidades (PQ)",
       "1 linha por item-folha da EAP. Disciplina e Frente/Trecho alimentam TODOS os rollups. Marque 'Sim' em Insumo FD? nos insumos de faturamento direto (cláusula de reequilíbrio).")
r0, r1 = tabela(ws,
    ["Nº item", "Código", "Descrição do serviço/insumo", "Unidade", "Qtde contratada",
     "Custo unit (R$)", "Custo total (R$)", "Preço venda unit c/ BDI (R$)", "Valor venda total (R$)",
     "Disciplina", "Frente/Trecho", "Grupo", "Insumo FD?"],
    [10, 14, 52, 9, 14, 14, 16, 16, 16, 22, 18, 14, 11],
    1500,
    fmts={5: FMT_INT, 6: FMT_RS, 7: FMT_RS, 8: FMT_RS, 9: FMT_RS},
    total_sum_cols=(7, 9))
dv_lista(ws, REF["Disciplina"], f"J{r0}:J{r1}")
dv_lista(ws, REF["Sim/Não"], f"M{r0}:M{r1}")

ws = aba(wb, "I-03 BDI", COR_IMUT)
titulo(ws, "I-03 · C.1 — BDI Detalhe (composição por rubrica)",
       "Composição do BDI da proposta. 'Rubrica de tempo?' = Sim para rubricas proporcionais ao prazo (Adm Central, seguros...) — alimenta o D.2.")
r0, r1 = tabela(ws,
    ["Rubrica", "% sobre receita", "% sobre custo direto", "Valor (R$)", "É subtotal?", "Rubrica de tempo?"],
    [38, 16, 18, 16, 12, 15],
    30,
    fmts={2: FMT_PCT, 3: FMT_PCT, 4: FMT_RS})
total_declarado(ws, r1 + 1, 4, FMT_RS, "SUM ingênuo dupla-contaria os subtotais — digite o BDI total oficial")
dv_lista(ws, REF["Sim/Não"], f"E{r0}:F{r1}")

ws = aba(wb, "I-04 Encargos", COR_IMUT)
titulo(ws, "I-04 · D.3 — Encargos Sociais (composição de alíquotas)",
       "Composição da proposta: grupos A/B/C/D + incidências. Usar coluna Grupo (não linha-subtítulo).")
r0, r1 = tabela(ws,
    ["Grupo", "Item / alíquota", "% ", "Base de incidência", "Observação"],
    [9, 44, 11, 30, 34],
    60,
    fmts={3: FMT_PCT})
total_declarado(ws, r1 + 1, 3, FMT_PCT, "incidências entre grupos (B sobre A etc.) impedem soma simples — digite o total oficial")

ws = aba(wb, "I-05 CPUs", COR_IMUT)
titulo(ws, "I-05 · CPU — Coeficientes (decomposição MOD/EQP por unidade de serviço)",
       "1 linha por composição de custo unitário. Invariante conferida pelo motor: MOD + EQP ≤ custo direto unit.")
tabela(ws,
    ["Código CPU", "Serviço", "Unidade", "Tipo", "Custo direto unit (R$)",
     "MOD (R$/un)", "EQP (R$/un)", "% MOD", "% EQP", "% Material"],
    [16, 52, 9, 14, 18, 14, 14, 10, 10, 11],
    1000,
    fmts={5: FMT_RS, 6: FMT_RS, 7: FMT_RS, 8: FMT_PCT, 9: FMT_PCT, 10: FMT_PCT})

ws = aba(wb, "I-06 Curva Prevista", COR_IMUT)
titulo(ws, "I-06 · C.3/C.5 — Curva prevista mensal (físico % e financeiro R$)",
       "1 linha por mês da obra. Σ % físico deve fechar 100% e Σ R$ deve fechar o PV — a linha TOTAL confere.")
tabela(ws,
    ["Mês", "Competência (mm/aaaa)", "% físico previsto no mês", "R$ previsto no mês"],
    [8, 20, 22, 20],
    N_MESES,
    fmts={1: FMT_INT, 3: FMT_PCT, 4: FMT_RS},
    total_sum_cols=(3, 4))

ws = aba(wb, "I-07 Previsto R$ Disc x Mês", COR_IMUT)
titulo(ws, "I-07 · C.3 — Faturamento previsto por Disciplina × Mês (R$)",
       "Matriz: 1 linha por disciplina, colunas = mês da obra (1..60). Σ da matriz deve bater a curva I-06 mês a mês e o PV no total.")
r0, r1 = matriz_meses(ws, 1, ["Disciplina"], [26], 20)
dv_lista(ws, REF["Disciplina"], f"A{r0}:A{r1}")

ws = aba(wb, "I-19 Previsto R$ Frente x Mês", COR_IMUT)
titulo(ws, "I-19 · C.3 — Faturamento previsto por Frente × Mês (R$)",
       "Matriz: 1 linha por frente/trecho (mesmos nomes da coluna Frente/Trecho da PQ), colunas = mês da obra. Baseline das curvas por frente (C.8) e do recorte por frente (C.3). Σ deve bater o PV.")
matriz_meses(ws, 1, ["Frente/Trecho"], [26], 25)

ws = aba(wb, "I-08 Previsto Fís Disc x Mês", COR_IMUT)
titulo(ws, "I-08 · C.5 — % físico previsto ACUMULADO por Disciplina × Mês",
       "Matriz: fração acumulada 0→1 por disciplina (monotônica — o motor confere). Última coluna preenchida = 100%.")
r0, r1 = matriz_meses(ws, 1, ["Disciplina"], [26], 20, fmt=FMT_PCT)
dv_lista(ws, REF["Disciplina"], f"A{r0}:A{r1}")

ws = aba(wb, "I-09 Gantt Tarefas", COR_IMUT)
titulo(ws, "I-09 · C.13 — Cronograma de tarefas (Gantt contratado)",
       "1 linha por tarefa. Coluna Grupo = trecho/frente da tarefa (não usar linha-subtítulo). Datas CONTRATADAS — as reais vão na M-05.")
r0, r1 = tabela(ws,
    ["Nº", "Nome da tarefa", "Grupo (trecho/frente)", "Duração (dias)", "Data início", "Data término", "É marco?"],
    [8, 48, 26, 14, 14, 14, 10],
    600,
    fmts={4: FMT_INT, 5: FMT_DATA, 6: FMT_DATA})
dv_lista(ws, REF["Sim/Não"], f"G{r0}:G{r1}")

ws = aba(wb, "I-10 Marcos", COR_IMUT)
titulo(ws, "I-10 · C.5 — Marcos contratuais detalhados",
       "1 linha por marco com data-limite contratual. O % concluído vai na M-06 (mensal).")
tabela(ws,
    ["Categoria", "Trecho/Frente", "Marco contratual", "Data-limite", "Observação"],
    [18, 20, 46, 14, 30],
    40,
    fmts={4: FMT_DATA})

ws = aba(wb, "I-11 Recursos Contratados", COR_IMUT)
titulo(ws, "I-11 · C.4 — Recursos contratados (catálogo MOD/MOI/EQP)",
       "1 linha por função/equipamento. 'Grupo Adm Local' só para MOI que compõe a Administração Local (alimenta o D.1 — não criar aba separada).")
r0, r1 = tabela(ws,
    ["Categoria", "Função / Equipamento", "Grupo Adm Local (só MOI)", "Qtde contratada",
     "Custo unit mensal (R$)", "Custo total (R$)", "Jornada (h/mês)",
     "Custo-hora produtivo (R$/h) — EQP", "Custo-hora improdutivo (R$/h) — EQP"],
    [12, 40, 26, 15, 18, 18, 14, 22, 24],
    150,
    fmts={4: FMT_INT, 5: FMT_RS, 6: FMT_RS, 7: FMT_INT, 8: FMT_RS, 9: FMT_RS},
    total_sum_cols=(6,))
dv_lista(ws, REF["Categoria"], f"A{r0}:A{r1}")

ws = aba(wb, "I-12 Histograma Contratado", COR_IMUT)
titulo(ws, "I-12 · C.4 — Histograma contratado (qtd de recursos por mês)",
       "Matriz: 1 linha por recurso (mesmos nomes da I-11), colunas = mês da obra. QUANTIDADES previstas — o R$ o motor deriva (qtd × custo unit da I-11).")
r0, r1 = matriz_meses(ws, 1, ["Categoria", "Recurso"], [12, 40], 150, fmt=FMT_INT)
dv_lista(ws, REF["Categoria"], f"A{r0}:A{r1}")

ws = aba(wb, "I-13 Fontes de Índice", COR_IMUT)
titulo(ws, "I-13 · C.6 — Fontes de índice dos insumos FD (valores na data da OS)",
       "Para cada insumo FD (marcado na I-02): 1 linha por fonte de índice disponível. Valor ATUAL (mensal) vai na M-12. Marcar 1 fonte recomendada ★ por insumo.")
r0, r1 = tabela(ws,
    ["Insumo (nome da PQ)", "Fonte", "Rótulo do seletor", "Código da fonte", "Tipo",
     "Valor na data da OS", "Recomendada?"],
    [42, 12, 22, 16, 10, 18, 14],
    300,
    fmts={6: FMT_RS})
dv_lista(ws, REF["Fonte de índice"], f"B{r0}:B{r1}")
dv_lista(ws, REF["Tipo de fonte"], f"E{r0}:E{r1}")
dv_lista(ws, REF["Sim/Não"], f"G{r0}:G{r1}")

ws = aba(wb, "I-14 Reequilíbrio Params", COR_IMUT)
titulo(ws, "I-14 · D.5 — Parâmetros de reajuste e reequilíbrio",
       "Datas e marcos das cláusulas de reajuste (M1) e reequilíbrio de insumos (M2). Números-índice mensais vão na M-11.")
chave_valor(ws, [
    ("Data da OS (marco do reequilíbrio)", FMT_DATA, ""),
    ("Data da proposta", FMT_DATA, "cenário M1 'proposta'"),
    ("Data da assinatura", FMT_DATA, "cenário M1 'assinatura'"),
    ("Data das CPUs (composições)", FMT_DATA, "cenário M1 'cpus' — se aplicável"),
    ("Data do reajuste de aniversário", FMT_DATA, ""),
    ("Data de verificação do reequilíbrio", FMT_DATA, "ex.: semestral"),
    ("Cláusula do reajuste geral (M1)", None, "ex.: 6.2"),
    ("Cláusula do reequilíbrio de insumos (M2)", None, "ex.: 8.8"),
    ("Teto/linha divisória do M2", None, "ex.: excedente sobre o índice contratual do período"),
])

ws = aba(wb, "I-15 Mapa Segmentos", COR_IMUT)
titulo(ws, "I-15 · C.14 — Mapa da obra (segmentos / frentes físicas)",
       "Obra LINEAR: 1 linha por segmento com km início/fim (geometria contígua — o motor confere). Obra NÃO-LINEAR (edificação, aeroporto…): 1 linha por frente física com km em branco. Liberação REAL vai na M-10.")
r0, r1 = tabela(ws,
    ["Código", "Nome do segmento", "Tipo", "km início", "km fim", "Valor contratado (R$)",
     "Mês de liberação prevista (nº)"],
    [10, 34, 16, 12, 12, 20, 24],
    60,
    fmts={4: "0.000", 5: "0.000", 6: FMT_RS, 7: FMT_INT},
    total_sum_cols=(6,))
dv_lista(ws, REF["Tipo de segmento"], f"C{r0}:C{r1}")

ws = aba(wb, "I-16 Mapa Elementos", COR_IMUT)
titulo(ws, "I-16 · C.14 — Elementos pontuais (OAE · dispositivos · taludes)",
       "1 linha por elemento pontual com localização. Valor obrigatório para taludes/sinistros.")
r0, r1 = tabela(ws,
    ["Tipo", "Elemento", "km", "Estaca", "Valor (R$)", "Lado / Observação"],
    [14, 40, 10, 12, 18, 30],
    80,
    fmts={3: "0.000", 5: FMT_RS},
    total_sum_cols=(5,))
dv_lista(ws, REF["Tipo de elemento"], f"A{r0}:A{r1}")

ws = aba(wb, "I-17 Chuvas Baseline", COR_IMUT)
titulo(ws, "I-17 · C.9 — Baseline de chuvas (média histórica)",
       "1 linha por mês da obra: chuva prevista pela série histórica da região + dias >5mm esperados. Real vai na M-07.")
tabela(ws,
    ["Mês", "Competência (mm/aaaa)", "Chuva prevista (mm)", "Dias >5mm previstos"],
    [8, 20, 18, 20],
    N_MESES,
    fmts={1: FMT_INT, 3: "0.0", 4: FMT_INT},
    total_sum_cols=(3, 4))

ws = aba(wb, "I-18 Parâmetros da Obra", COR_IMUT)
titulo(ws, "I-18 · Config — Parâmetros e metas da obra",
       "Metas internas e parâmetros de cálculo (não vêm do contrato). Alimentam C.7, D.6 e as réguas de farol.")
chave_valor(ws, [
    ("Meta de produtividade (R$/HH)", FMT_RS, "meta do projeto"),
    ("Jornada MOD (h/mês)", FMT_INT, ""),
    ("Jornada MOI (h/mês)", FMT_INT, ""),
    ("Jornada diária (h/dia)", "0.0", "para cálculo de ociosidade D.6"),
    ("Custo-hora MOD (R$/h)", FMT_RS, "para eventos/ociosidade D.6"),
    ("Custo-hora EQP (R$/h)", FMT_RS, ""),
    ("Benchmark produtividade — fonte", None, "ex.: TCPO/SINAPI"),
    ("Benchmark produtividade — valor (R$/HH)", FMT_RS, ""),
    ("Régua de farol custom (se houver)", None, "descrever; padrão da plataforma se em branco"),
])

# ══════════════ PARTE 2 · MUTÁVEIS ═══════════════════════════════════════════

ws = aba(wb, "M-01 Capa do Ciclo", COR_MUT)
titulo(ws, "M-01 · BM — Registro dos ciclos (1 linha por BM)",
       "1 linha por BM, ACUMULA a obra inteira (nunca apagar ciclo anterior). O total declarado é a âncora do gate da medição (M-02).", mut=True)
tabela(ws,
    ["BM nº", "Competência (mm/aaaa)", "Data de corte", "Valor total medido no período (R$) — declarado",
     "Responsável (cargo/função)"],
    [8, 20, 14, 34, 26],
    60,
    fmts={1: FMT_INT, 3: FMT_DATA, 4: FMT_RS},
    mut=True)

ws = aba(wb, "M-02 Medição BM", COR_MUT)
titulo(ws, "M-02 · C.3 — Medição por período (Boletim de Medição)",
       "1 linha por item medido EM CADA BM — os ciclos ACUMULAM nesta aba (nunca apagar/sobrescrever BMs anteriores). O motor deriva daqui: faturamento real, avanço físico, VA (D.4) e insumos medidos (C.6). Acumulados NUNCA se digitam.", mut=True)
tabela(ws,
    ["BM nº (= M-01)", "Nº item (= I-02 PQ)", "Qtd medida no período", "Glosa (R$)", "Motivo da glosa / Observação"],
    [12, 18, 20, 14, 44],
    6000,
    fmts={1: FMT_INT, 3: FMT_INT, 4: FMT_RS},
    total_sum_cols=(4,),
    mut=True)

ws = aba(wb, "M-03 Recursos Reais", COR_MUT)
titulo(ws, "M-03 · C.4 — Recursos reais (por recurso × mês · formato longo)",
       "1 linha por recurso × mês: quantidade efetiva, HH trabalhadas e custo real. Alimenta C.4, C.7 (HH), D.1 (Adm Local real) e D.3. Meses sem apontamento ficam SEM linha (não zerar).", mut=True)
r0, r1 = tabela(ws,
    ["Mês (nº)", "Competência (mm/aaaa)", "Categoria", "Recurso (= I-11)", "Qtd real",
     "HH real", "Custo real no mês (R$)"],
    [9, 18, 12, 40, 11, 12, 20],
    3000,
    fmts={1: FMT_INT, 5: FMT_INT, 6: FMT_INT, 7: FMT_RS},
    total_sum_cols=(7,),
    mut=True)
dv_lista(ws, REF["Categoria"], f"C{r0}:C{r1}")

ws = aba(wb, "M-04 Avanço Físico Real", COR_MUT)
titulo(ws, "M-04 · C.5 — Avanço físico real (por disciplina × mês · acumulado)",
       "1 linha por disciplina × mês com o % físico real ACUMULADO (fração 0→1, monotônico). Fonte oficial: medição do BM.", mut=True)
r0, r1 = tabela(ws,
    ["Mês (nº)", "Competência (mm/aaaa)", "Disciplina", "% físico real acumulado"],
    [9, 18, 26, 22],
    1200,
    fmts={1: FMT_INT, 4: FMT_PCT},
    mut=True)
dv_lista(ws, REF["Disciplina"], f"C{r0}:C{r1}")

ws = aba(wb, "M-05 Gantt Real", COR_MUT)
titulo(ws, "M-05 · C.13 — Eixo real do cronograma (por tarefa)",
       "Atualizar a tarefa quando INICIAR (data início real) e quando TERMINAR (data término real). Tarefa em andamento: término em branco + % concluído.", mut=True)
tabela(ws,
    ["Nº tarefa (= I-09)", "Data início real", "Data término real", "% concluído",
     "Competência do registro (mm/aaaa)", "Observação"],
    [16, 15, 16, 12, 24, 36],
    150,
    fmts={2: FMT_DATA, 3: FMT_DATA, 4: FMT_PCT},
    mut=True)

ws = aba(wb, "M-06 Marcos Real", COR_MUT)
titulo(ws, "M-06 · C.5 — Situação dos marcos contratuais",
       "% concluído por marco (= I-10) no corte do BM. Marco não iniciado fica em branco.", mut=True)
tabela(ws,
    ["Marco (= I-10)", "Competência (mm/aaaa)", "% concluído", "Observação"],
    [46, 20, 13, 36],
    120,
    fmts={3: FMT_PCT},
    mut=True)

ws = aba(wb, "M-07 Chuvas do Mês", COR_MUT)
titulo(ws, "M-07 · C.9 — Chuvas realizadas (por mês)",
       "1 linha por mês medido: mm real, dias >5mm e dias parados. Meses futuros ficam SEM linha.", mut=True)
tabela(ws,
    ["Mês (nº)", "Competência (mm/aaaa)", "Chuva real (mm)", "Dias >5mm reais", "Dias parados"],
    [9, 18, 15, 16, 13],
    N_MESES,
    fmts={1: FMT_INT, 3: "0.0", 4: FMT_INT, 5: FMT_INT},
    mut=True)

ws = aba(wb, "M-08 Chuva Diária RDO", COR_MUT)
titulo(ws, "M-08 · D.6 — Chuva diária do RDO (OPCIONAL · habilita 'dias a cobrar')",
       "1 linha por dia com registro no RDO — inclusive 0 mm (série completa vale mais). Habilita a apuração diária de ociosidade da D.6 e o headline da C.9.", mut=True)
tabela(ws,
    ["Data", "Chuva (mm)", "Efetivo MOD em campo", "Equipamentos em produção", "Observação do RDO"],
    [12, 11, 20, 22, 40],
    2000,
    fmts={1: FMT_DATA, 2: "0.0", 3: FMT_INT, 4: FMT_INT},
    mut=True)

ws = aba(wb, "M-09 Eventos e Impedimentos", COR_MUT)
titulo(ws, "M-09 · C.13/D.6 — Registro de eventos que impactam prazo/custo",
       "Cadastro INCREMENTAL: nunca apagar linha; evento encerrado ganha data fim. 1 linha por evento, com fonte documental (RDO/ata/CE/ofício).", mut=True)
r0, r1 = tabela(ws,
    ["ID", "Data início", "Data fim", "Título do evento (fato real)", "Categoria",
     "Frente/Trecho/km", "Dias", "MOD afetado", "EQP afetado", "Crítico?",
     "Cláusulas relacionadas", "Fonte (RDO/ata/CE...)", "Observação"],
    [8, 13, 13, 44, 15, 18, 8, 12, 12, 10, 20, 20, 30],
    200,
    fmts={2: FMT_DATA, 3: FMT_DATA, 7: FMT_INT, 8: FMT_INT, 9: FMT_INT},
    mut=True)
dv_lista(ws, REF["Categoria de evento"], f"E{r0}:E{r1}")
dv_lista(ws, REF["Sim/Não"], f"J{r0}:J{r1}")

ws = aba(wb, "M-10 Liberação de Frentes", COR_MUT)
titulo(ws, "M-10 · C.14 — Liberação e impedimento de frentes (por segmento)",
       "1 linha por segmento (= I-15) com o eixo real: mês de liberação real e janelas de impedimento. Segmento ainda não liberado fica em branco.", mut=True)
tabela(ws,
    ["Segmento (= I-15)", "Mês liberação real (nº)", "Impedimento início (mês nº)",
     "Impedimento fim (mês nº)", "Causa do impedimento"],
    [26, 20, 22, 20, 40],
    60,
    fmts={2: FMT_INT, 3: FMT_INT, 4: FMT_INT},
    mut=True)

ws = aba(wb, "M-11 Série Índice Contratual", COR_MUT)
titulo(ws, "M-11 · D.5 — Série do índice contratual (número-índice)",
       "1 linha por mês publicado (append mensal). Marcar o Cenário nas linhas das datas-base do M1 (cpus/proposta/assinatura).", mut=True)
r0, r1 = tabela(ws,
    ["Competência (mm/aaaa)", "Número-índice", "Cenário M1", "Observação"],
    [20, 16, 14, 34],
    120,
    fmts={2: "#,##0.00"},
    mut=True)
dv_lista(ws, REF["Cenário M1"], f"C{r0}:C{r1}")

ws = aba(wb, "M-12 Índices dos Insumos", COR_MUT)
titulo(ws, "M-12 · C.6 — Índices dos insumos por competência (série · append)",
       "1 linha por competência × insumo × fonte — ACUMULA a série mensal (nunca sobrescrever mês anterior). O Δ% vs OS e o excedente o motor deriva; a série alimenta os gráficos de evolução.", mut=True)
r0, r1 = tabela(ws,
    ["Competência (mm/aaaa)", "Insumo (= I-13)", "Fonte", "Valor do índice no mês", "Observação"],
    [20, 42, 12, 18, 28],
    2000,
    fmts={4: FMT_RS},
    mut=True)
dv_lista(ws, REF["Fonte de índice"], f"C{r0}:C{r1}")

ws = aba(wb, "M-13 Custos Pagos", COR_MUT)
titulo(ws, "M-13 · C.6 — Custos pagos (OPCIONAL · preço real de insumos)",
       "1 linha por NF relevante de insumo: habilita o preço real pago (C.6) e o custo real (C.7). Sem NF, deixar aba vazia.", mut=True)
tabela(ws,
    ["Data", "NF", "Insumo / Recurso", "Qtd", "Preço unit pago (R$)", "Valor (R$)", "Observação"],
    [12, 14, 40, 10, 18, 16, 30],
    1000,
    fmts={1: FMT_DATA, 4: FMT_INT, 5: FMT_RS, 6: FMT_RS},
    total_sum_cols=(6,),
    mut=True)

# ordem das abas meta: LEIA-ME · Guia da IA · LISTAS
wb.move_sheet("Guia da IA", offset=-1)

# ── salvar ────────────────────────────────────────────────────────────────────
SAIDA.parent.mkdir(parents=True, exist_ok=True)
wb.save(SAIDA)
n_abas = len(wb.sheetnames)
print(f"✅ {SAIDA.relative_to(RAIZ)} gerada — {n_abas} abas:")
for s in wb.sheetnames:
    print("  ·", s)
